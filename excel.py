from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, DEFAULT_FONT, Border, Side
from openpyxl.formatting.rule import CellIsRule

def write_excel(year, scores, excel, new_standards, old_standards):
    subjects = ['國', '英', '數A', '數B', '社', '自']

    subject_set = set()
    for row in excel:
        for subject in row[5]:
            subject_set.add(subject)
    subject_set = sorted(
        subject_set,
        key=lambda x: (len(x.split('+')), subjects.index(x.split('+')[0]))
    )

    wb = Workbook()
    wb.remove(wb.active)

    # 校系錄取標準
    wb.create_sheet(f'{year-1}校系錄取標準')
    sheet = wb[f'{year-1}校系錄取標準']

    # 表頭
    sheet['A1'] = '校系名稱'
    sheet.merge_cells('A1:A2')
    sheet['B1'] = '校系代碼'
    sheet.merge_cells('B1:B2')
    sheet['C1'] = '招生名額'
    sheet.merge_cells('C1:C2')
    sheet['D1'] = f'{year-1}分科'
    sheet.merge_cells('D1:D2')

    start = 5
    for i, subject in enumerate([*subjects, '英聽']):
        col_number = start + i * 2
        left = get_column_letter(col_number)
        right = get_column_letter(col_number+1)
        sheet[f'{left}2'] = subject
        sheet.merge_cells(f'{left}2:{right}2')
    col_number += 2
    sheet.cell(2, col_number).value = 'Pass?'
    end = col_number
    left = get_column_letter(start)
    right = get_column_letter(end)
    sheet[f'{left}1'] = '檢定標準'
    sheet.merge_cells(f'{left}1:{right}1')

    start = end + 1
    for i, subject in enumerate([*subjects, '組合']):
        col_number = start + i
        sheet.cell(2, col_number).value = subject
    end = col_number
    left = get_column_letter(start)
    right = get_column_letter(end)
    sheet[f'{left}1'] = '篩選倍率'
    sheet.merge_cells(f'{left}1:{right}1')

    start = end + 1
    for i, subject in enumerate(subject_set):
        col_number = start + i * 3
        sheet.cell(2, col_number).value = ''.join(subject.split('+'))
        sheet.cell(2, col_number+1).value = '原始'
        sheet.cell(2, col_number+2).value = '調整'
    col_number += 3
    sheet.cell(2, col_number).value = '原始'
    sheet.cell(2, col_number+1).value = '調整'
    end = col_number + 1
    left = get_column_letter(start)
    right = get_column_letter(end)
    sheet[f'{left}1'] = '通過篩選分數'
    sheet.merge_cells(f'{left}1:{right}1')

    start = end + 1
    for i, title in enumerate(['審查資料外指定項目', '寄發甄試通知', '繳交資料截止', '甄試日期']):
        col_number = start + i
        col = get_column_letter(col_number)
        sheet[f'{col}1'] = title
        sheet.merge_cells(f'{col}1:{col}2')
    
    # 內容
    subject_cols = {
        subject: chr(col)
        for subject, col in zip(
            [*subjects, '英聽'],
            range(ord('B'), ord('H')+1)
        )
    }

    for i, (school_name, department_name, department_id, info, average, screen_result) in enumerate(excel):
        row_number = 3 + i
        # 基本資料
        row = [school_name + department_name, int(department_id), int(info['number']), average]
        # 檢定標準
        start = 5
        for j, subject in enumerate([*subjects, '英聽']):
            score_col_number = start + j * 2
            standard = info['subjects'].get(subject)
            if standard is None:
                row.append('')
            else:
                standard = standard[0]
                if standard == '--':
                    row.append('')
                else:
                    row.append(standard)
            score_col = get_column_letter(score_col_number)
            score_cell = f'{score_col}{row_number}'
            subject_col = subject_cols[subject]
            if subject == '英聽':
                formula = f'=IF({score_cell}="","",IF(AND({score_cell}="A",學測分數!{subject_col}$2="A"),"V",IF(AND({score_cell}="B",OR(學測分數!{subject_col}$2="A",學測分數!{subject_col}$2="B")),"V",IF(AND({score_cell}="C",OR(學測分數!{subject_col}$2="A",學測分數!{subject_col}$2="B",學測分數!{subject_col}$2="C")),"V","X"))))'
            else:
                formula = f'=IF({score_cell}="","",IF(AND({score_cell}="頂標",學測分數!{subject_col}$2>=學測分數!{subject_col}$4),"V",IF(AND({score_cell}="前標",學測分數!{subject_col}$2>=學測分數!{subject_col}$5),"V",IF(AND({score_cell}="均標",學測分數!{subject_col}$2>=學測分數!{subject_col}$6),"V",IF(AND({score_cell}="後標",學測分數!{subject_col}$2>=學測分數!{subject_col}$7),"V",IF(AND({score_cell}="底標",學測分數!{subject_col}$2>=學測分數!{subject_col}$8),"V","X"))))))'
            row.append(formula)
        end = score_col_number + 1
        left = get_column_letter(start)
        right = get_column_letter(end)
        formula = f'=IF(COUNTIF({left}{row_number}:{right}{row_number},"X")>0,"L","J")'
        row.append(formula)
        # 篩選倍率
        for subject in [*subjects, '組合']:
            rate = info['subjects'].get(subject)
            if rate is None:
                row.append('')
            else:
                rate = rate[1]
                if rate == '--':
                    row.append('')
                else:
                    row.append(float(rate) if '.' in rate else int(rate))
        # 通過篩選分數
        start = end + len(subjects) + 3
        origin_col_numbers = []
        adjust_col_numbers = []
        for j, subject in enumerate(subject_set):
            score_col_number = start + j * 3
            origin_col_numbers.append(score_col_number+1)
            adjust_col_numbers.append(score_col_number+2)
            score_col = get_column_letter(score_col_number)
            score_cell = f'{score_col}{row_number}'
            score = screen_result.get(subject)
            if score is None:
                row.append('')
            else:
                row.append(float(score) if '.' in score else int(score))
            expressions = [
                f'學測分數!{subject_cols[s]}$2'
                for s in subject.split('+')
            ]
            expression = '+'.join(expressions)
            formula = f'=IF({score_cell}="","",IF({expression}-{score_cell}>=0,"J",IF({expression}-{score_cell}>=-1,"K","L")))'
            row.append(formula)
            expressions = [
                f'學測分數!{subject_cols[s]}$3'
                for s in subject.split('+')
            ]
            expression = '+'.join(expressions)
            formula = f'=IF({score_cell}="","",IF({expression}-{score_cell}>=0,"J",IF({expression}-{score_cell}>=-1,"K","L")))'
            row.append(formula)
        formula_L = []
        formula_K = []
        for col_number in origin_col_numbers:
            col = get_column_letter(col_number)
            cell = f'{col}{row_number}'
            formula_L.append(f'{cell}="L"')
            formula_K.append(f'{cell}="K"')
        or_L = f'OR({",".join(formula_L)})'
        or_K = f'OR({",".join(formula_K)})'
        formula = f'=IF({or_L},"L",IF({or_K},"K","J"))'
        row.append(formula)
        formula_L = []
        formula_K = []
        for col_number in adjust_col_numbers:
            col = get_column_letter(col_number)
            cell = f'{col}{row_number}'
            formula_L.append(f'{cell}="L"')
            formula_K.append(f'{cell}="K"')
        or_L = f'OR({",".join(formula_L)})'
        or_K = f'OR({",".join(formula_K)})'
        formula = f'=IF({or_L},"L",IF({or_K},"K","J"))'
        row.append(formula)
        # 其他資訊
        row.append('、'.join(info['meetings']))
        row.append(info['send_time'])
        row.append(info['end_time'])
        row.append(info['meeting_time'])

        sheet.append(row)

    max_row_number = sheet.max_row
    max_col_number = sheet.max_column
    max_col = get_column_letter(max_col_number)

    # 字體大小
    DEFAULT_FONT.size = 12

    # 粗體
    bold_14 = Font(bold='bold', size=14)
    def set_func(cell, font):
        if cell.value not in ['原始', '調整']:
            cell.font = font
    set_range_style(sheet, f'A1:{max_col}2', set_func, bold_14)
    col_number = max_col_number - 5
    sheet.cell(2, col_number).font = bold_14
    sheet.cell(2, col_number+1).font = bold_14

    # 特殊字型
    wingdings = Font('Wingdings')
    def set_func(cell, font):
        cell.font = font
    col_number = 4 + len(subjects) * 2 + 3
    col = get_column_letter(col_number)
    set_range_style(sheet, f'{col}3:{col}{max_row_number}', set_func, wingdings)
    col_number += len(subjects) + 2
    for i in range(len(subject_set)):
        start = col_number + i * 3 + 1
        left = get_column_letter(start)
        right = get_column_letter(start+1)
        set_range_style(sheet, f'{left}3:{right}{max_row_number}', set_func, wingdings)
    left = get_column_letter(start+2)
    right = get_column_letter(start+3)
    set_range_style(sheet, f'{left}3:{right}{max_row_number}', set_func, wingdings)

    # 置中
    center = Alignment('center', 'center')
    def set_func(cell, alignment):
        cell.alignment = alignment
    set_range_style(sheet, f'A1:{max_col}2', set_func, center)
    col_number = max_col_number - 4
    col = get_column_letter(col_number)
    set_range_style(sheet, f'B3:{col}{max_row_number}', set_func, center)

    # 數值格式
    def set_func(cell):
        cell.number_format = '000000'
    set_range_style(sheet, f'B3:B{max_row_number}', set_func)

    # 填充顏色
    fill = PatternFill('solid', fgColor='D9D9D9')
    def set_func(cell, fill):
        cell.fill = fill
    set_range_style(sheet, f'A1:{max_col}2', set_func, fill)
    fill = PatternFill('solid', fgColor='FFFF00')
    set_range_style(sheet, f'B3:B{max_row_number}', set_func, fill)
    fill = PatternFill('solid', fgColor='FCE4D6')
    col_number = 5 + len(subjects) * 2 + 2
    col = get_column_letter(col_number)
    set_range_style(sheet, f'E3:{col}{max_row_number}', set_func, fill)
    fill = PatternFill('solid', fgColor='DDEBF7')
    start = col_number + 1
    end = start + len(subjects)
    left = get_column_letter(start)
    right = get_column_letter(end)
    set_range_style(sheet, f'{left}3:{right}{max_row_number}', set_func, fill)
    fill = PatternFill('solid', fgColor='E2EFDA')
    start = end + 1
    end = start + len(subject_set) * 3 + 1
    left = get_column_letter(start)
    right = get_column_letter(end)
    set_range_style(sheet, f'{left}3:{right}{max_row_number}', set_func, fill)

    # 邊框
    border = Border(Side('thin'), Side('thin'), Side('thin'), Side('thin'))
    def set_func(cell, border):
        cell.border = border
    set_range_style(sheet, f'A1:{max_col}{max_row_number}', set_func, border)

    # 格式化條件
    red = Font(color='9C0006')
    yellow = Font(color='ED7D31')
    green = Font(color='00B050')
    blue = Font(color='4472C4')
    rules = [
        CellIsRule('equal', ['"頂標"'], font=red),
        CellIsRule('equal', ['"前標"'], font=yellow),
        CellIsRule('equal', ['"均標"'], font=green),
        CellIsRule('equal', ['"後標"'], font=blue)
    ]
    col_number = 5
    for i in range(len(subjects)):
        start = col_number + i * 2
        col = get_column_letter(start)
        for rule in rules:
            sheet.conditional_formatting.add(f'{col}3:{col}{max_row_number}', rule)
        col = get_column_letter(start+1)
        sheet.column_dimensions[col].hidden = True
    rules = [
        CellIsRule('equal', ['"A"'], font=red),
        CellIsRule('equal', ['"B"'], font=green)
    ]
    col_number = start + 2
    col = get_column_letter(col_number)
    for rule in rules:
        sheet.conditional_formatting.add(f'{col}3:{col}{max_row_number}', rule)
    col = get_column_letter(col_number+1)
    sheet.column_dimensions[col].hidden = True
    red = Font(color='FF0000')
    yellow = Font(color='FFC000')
    rules = [
        CellIsRule('equal', ['"L"'], font=red),
        CellIsRule('equal', ['"K"'], font=yellow),
        CellIsRule('equal', ['"J"'], font=green)
    ]
    col_number += 2
    col = get_column_letter(col_number)
    for rule in rules:
        sheet.conditional_formatting.add(f'{col}3:{col}{max_row_number}', rule)
    col_number += len(subjects) + 2
    for i in range(len(subject_set)):
        start = col_number + i * 3 + 1
        col = get_column_letter(start)
        for rule in rules:
            sheet.conditional_formatting.add(f'{col}3:{col}{max_row_number}', rule)
        sheet.column_dimensions[col].hidden = True
        col = get_column_letter(start+1)
        for rule in rules:
            sheet.conditional_formatting.add(f'{col}3:{col}{max_row_number}', rule)
        sheet.column_dimensions[col].hidden = True
    col_number = start + 2
    col = get_column_letter(col_number)
    for rule in rules:
        sheet.conditional_formatting.add(f'{col}3:{col}{max_row_number}', rule)
    col = get_column_letter(col_number+1)
    for rule in rules:
        sheet.conditional_formatting.add(f'{col}3:{col}{max_row_number}', rule)

    # 調整寬度
    length = max([
        0 if cell.value is None else len(cell.value)
        for cell in sheet['A']
    ])
    sheet.column_dimensions['A'].width = get_width(length)
    sheet.column_dimensions['B'].width = get_width(4, 0, 14)
    sheet.column_dimensions['C'].width = get_width(4, 0, 14)
    sheet.column_dimensions['D'].width = get_width(2, 3, 14)
    start = 5
    for i, subject in enumerate([*subjects, '英聽']):
        col_number = start + i * 2
        col = get_column_letter(col_number)
        size = 14 if subject in ['數A', '數B', '英聽'] else 12
        extra = 4 if subject == '英聽' else 2 if subject[0] == '數' else 1
        sheet.column_dimensions[col].width = get_width(2, 0, size) + extra
        col = get_column_letter(col_number+1)
        sheet.column_dimensions[col].width = get_width(0, 1)
    col_number += 2
    col = get_column_letter(col_number)
    sheet.column_dimensions[col].width = get_width(1, 4, 14) + 1
    start = col_number + 1
    for i, subject in enumerate([*subjects, '組合']):
        col_number = start + i
        col = get_column_letter(col_number)
        extra = 2 if subject[0] == '數' else 3
        sheet.column_dimensions[col].width = get_width(len(subject), 0, 14) + extra
    start = col_number + 1
    for i, subject in enumerate(subject_set):
        col_number = start + i * 3
        col = get_column_letter(col_number)
        subject = ''.join(subject.split('+'))
        extra = 2 if '數' in subject else 3
        sheet.column_dimensions[col].width = get_width(len(subject), 0, 14) + extra
        col = get_column_letter(col_number+1)
        sheet.column_dimensions[col].width = get_width(2) + 3 / 14 * 12
        col = get_column_letter(col_number+2)
        sheet.column_dimensions[col].width = get_width(2) + 3 / 14 * 12
    col_number += 3
    col = get_column_letter(col_number)
    sheet.column_dimensions[col].width = get_width(2, 0, 14) + 3
    col = get_column_letter(col_number+1)
    sheet.column_dimensions[col].width = get_width(2, 0, 14) + 3
    col_number += 2
    col = get_column_letter(col_number)
    widths = []
    for i, cell in enumerate(sheet[col]):
        if cell.value is None:
            widths.append(0)
            continue
        size = 14 if i < 2 else 12
        widths.append(get_width(len(cell.value), 0, size))
    sheet.column_dimensions[col].width = max(widths)
    for i in range(3):
        col = get_column_letter(col_number+i+1)
        widths = []
        for i, cell in enumerate(sheet[col]):
            if cell.value is None:
                widths.append(0)
                continue
            size = 14 if i < 2 else 12
            ch = 0
            en = 0
            for text in cell.value:
                if 0x4e00 <= ord(text) <= 0x9fff:
                    ch += 1
                elif '0' <= text <= '9':
                    en += 1
                else:
                    en += 0.25
            widths.append(get_width(ch, en, size))
        sheet.column_dimensions[col].width = max(widths)

    # 凍結窗格
    sheet.freeze_panes = 'B3'

    # 篩選
    sheet.auto_filter.ref = f'A2:{max_col}2'

    # 學測分數
    wb.create_sheet('學測分數')
    sheet = wb['學測分數']

    sheet.append([year, *subjects, '英聽'])
    sheet.append(['原始分數', *scores])
    sheet.append([
        '調整分數',
        *[
            f'=ROUND(IF({chr(col)}2>={chr(col)}4,({chr(col)}2-{chr(col)}4)/(15-{chr(col)}4)*(15-{chr(col)}12)+{chr(col)}12,IF({chr(col)}2>={chr(col)}5,({chr(col)}2-{chr(col)}5)/({chr(col)}4-{chr(col)}5)*({chr(col)}12-{chr(col)}13)+{chr(col)}13,IF({chr(col)}2>={chr(col)}6,({chr(col)}2-{chr(col)}6)/({chr(col)}5-{chr(col)}6)*({chr(col)}13-{chr(col)}14)+{chr(col)}14,IF({chr(col)}2>={chr(col)}7,({chr(col)}2-{chr(col)}7)/({chr(col)}6-{chr(col)}7)*({chr(col)}14-{chr(col)}15)+{chr(col)}15,IF({chr(col)}2>={chr(col)}8,({chr(col)}2-{chr(col)}8)/({chr(col)}7-{chr(col)}8)*({chr(col)}15-{chr(col)}16)+{chr(col)}16,{chr(col)}2/{chr(col)}8*{chr(col)}16))))),0)'
            for col in range(ord('B'), ord('G')+1)
        ]
    ])
    # 今年各標
    for i, text in enumerate('頂前均後底'):
        sheet.append([
            f'{text}標',
            *[
                standard[i]
                for standard in new_standards.values()
            ]
        ])
    
    sheet.append([])
    sheet.append([])

    sheet.append([year-1, *subjects])
    for i, text in enumerate('頂前均後底'):
        sheet.append([
            f'{text}標',
            *[
                standard[i]
                for standard in old_standards.values()
            ]
        ])
    
    wb.save(f'{year-1}.xlsx')

def set_range_style(sheet, range, set_func, *args):
    for row in sheet[range]:
        for cell in row:
            set_func(cell, *args)

def get_width(ch=0, en=0, size=12):
    return (ch * 2 + en) * size / 11 + 2